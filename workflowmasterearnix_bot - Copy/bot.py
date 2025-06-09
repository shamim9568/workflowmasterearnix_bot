import os
import requests
from bs4 import BeautifulSoup
from pyrogram import Client, filters
from pyrogram.types import InlineKeyboardMarkup, InlineKeyboardButton
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

API_ID = 21969749
API_HASH = "e3c85629554abdf3bcea182fff83070b"
BOT_TOKEN = "8157106293:AAEZ2gy9Y1_8FwYjB4hnHqqTEQuuacR-Rs"
ADMIN_ID = 6822761216
SUPPORT_USERNAME = "@Shmim79"  # Telegram username or user id for support

app = Client("workflowmasterearnix_bot", bot_token=BOT_TOKEN, api_id=API_ID, api_hash=API_HASH)

user_wallets = {}
user_bkash = {}
pending_withdrawals = {}

@app.on_message(filters.command("start"))
def start(client, message):
    user = message.from_user
    user_wallets.setdefault(user.id, 0.0)

    keyboard = InlineKeyboardMarkup([
        [InlineKeyboardButton("\U0001F464 Account", callback_data="account")],
        [InlineKeyboardButton("\U0001F4E4 Submit File", callback_data="submit")],
        [InlineKeyboardButton("\U0001F4BC Wallet", callback_data="wallet")],
        [InlineKeyboardButton("\U0001F4B8 Withdraw", callback_data="withdraw")],
        [InlineKeyboardButton("\U0001F198 Support", url=f"https://t.me/{SUPPORT_USERNAME}")]
    ])

    message.reply_text(f"\U0001F31F Welcome {user.first_name}!\nSelect an option below:", reply_markup=keyboard)

@app.on_callback_query()
def handle_callback(client, callback_query):
    user = callback_query.from_user
    data = callback_query.data

    if data == "account":
        bkash = user_bkash.get(user.id, "Not Set")
        callback_query.message.reply_text(f"\U0001F464 Account Info:\nUser ID: {user.id}\nbKash: {bkash}")

    elif data == "submit":
        callback_query.message.reply_text("\U0001F4E4 Please send your XLSX file now.")

    elif data == "wallet":
        balance = user_wallets.get(user.id, 0.0)
        bkash = user_bkash.get(user.id, "Not Set")
        callback_query.message.reply_text(f"\U0001F4B0 Wallet Balance: {balance:.2f} BDT\nbKash: {bkash}")

    elif data == "withdraw":
        client.send_message(user.id, "/withdraw")

@app.on_message(filters.command("setbkash"))
def set_bkash(client, message):
    args = message.text.split()
    if len(args) < 2:
        return message.reply_text("\u2757 Please provide your bKash number.\nExample: /setbkash 01XXXXXXXXX")
    user_bkash[message.from_user.id] = args[1]
    message.reply_text("\u2705 bKash number set successfully.")

@app.on_message(filters.command("wallet"))
def check_wallet(client, message):
    uid = message.from_user.id
    balance = user_wallets.get(uid, 0.0)
    bkash = user_bkash.get(uid, "Not Set")
    message.reply_text(f"\U0001F4B0 Wallet Balance: {balance:.2f} BDT\nbKash: {bkash}")

@app.on_message(filters.command("submit") & filters.private)
def submit_work(client, message):
    message.reply_text("\U0001F4E4 Please send your XLSX file now.")

@app.on_message(filters.document & filters.private)
def handle_document(client, message):
    if not message.document.file_name.endswith(".xlsx"):
        return message.reply_text("\u274C Please upload an XLSX file.")

    file_path = app.download_media(message.document)
    wb = load_workbook(filename=file_path)
    sheet = wb.active

    usernames = [str(cell.value).strip() for row in sheet.iter_rows(min_row=2, max_col=1) for cell in row if cell.value]
    total = len(usernames)

    uid = message.from_user.id

    message.reply_text(
        f"""✅ File received.
Total usernames: {total}
💸 Earnings will be credited once reviewed by admin."""
    )

    app.send_document(
        chat_id=ADMIN_ID,
        document=file_path,
        caption=f"\U0001F4E4 New Work Submitted\nUser: {message.from_user.mention}\nTotal Accounts: {total}\nUse /credit {uid} <amount> to add balance."
    )

    os.remove(file_path)

@app.on_message(filters.command("withdraw"))
def request_withdraw(client, message):
    uid = message.from_user.id
    balance = user_wallets.get(uid, 0.0)
    if balance < 100:
        return message.reply_text("\u274C You need at least 100 BDT to request a withdrawal.")

    bkash = user_bkash.get(uid)
    if not bkash:
        return message.reply_text("\u26A0\uFE0F You need to set your bKash number first using /setbkash")

    pending_withdrawals[uid] = balance
    user_wallets[uid] = 0.0

    message.reply_text("\u2705 Withdrawal request sent to admin.")

    client.send_message(
        ADMIN_ID,
        f"\U0001F4B3 Withdrawal Request\nUser: {message.from_user.mention} (ID: {uid})\nbKash: {bkash}\nAmount: {balance:.2f} BDT\n\nTo confirm, use /pay {uid} amount"
    )

@app.on_message(filters.command("pay") & filters.user(ADMIN_ID))
def confirm_payment(client, message):
    args = message.text.split()
    if len(args) != 3:
        return message.reply_text("Usage: /pay <user_id> <amount>")

    try:
        uid = int(args[1])
        amount = float(args[2])
    except ValueError:
        return message.reply_text("Invalid format.")

    pending_withdrawals.pop(uid, None)
    client.send_message(uid, f"\u2705 Admin confirmed payment of {amount:.2f} BDT to your bKash. Thank you!")

@app.on_message(filters.command("credit") & filters.user(ADMIN_ID))
def credit_balance(client, message):
    args = message.text.split()
    if len(args) != 3:
        return message.reply_text("Usage: /credit <user_id> <amount>")

    try:
        uid = int(args[1])
        amount = float(args[2])
    except ValueError:
        return message.reply_text("Invalid user ID or amount.")

    user_wallets[uid] = user_wallets.get(uid, 0.0) + amount
    client.send_message(uid, f"\U0001F4B0 Admin credited {amount:.2f} BDT to your wallet.")
    message.reply_text(f"✅ {amount:.2f} BDT credited to user {uid}.")

app.run()





